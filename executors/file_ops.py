"""File-system + File Explorer executor.

Two concerns living together because they share the same data — local paths,
network shares, and the explorer.exe window:

1. **Programmatic FS ops** (read_excel, read_pdf, copy_to_local,
   find_latest_file, acquire_lock / release_lock). Cross-platform. Used by
   the extraction pipeline in Phase 4.
2. **File Explorer driving** (file_navigate / file_open via explorer.exe +
   pywinauto address bar). Windows-only at runtime; routed through
   `DesktopExecutor` when one is provided.

Phase 3 scope: implement (1) so the rest of the system has reliable file
plumbing, and (2) so the ActionRouter can route `file_navigate` / `file_open`
to a real explorer window.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import TYPE_CHECKING

from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_fixed

from agent.schemas import ActionPlan, ActionResult
from config.logging_config import get_logger
from config.settings import settings

if TYPE_CHECKING:
    from executors.desktop import DesktopExecutor

log = get_logger(__name__)


class FileError(RuntimeError):
    """Raised on file-system or File Explorer failures."""


class FileExecutor:
    EXPLORER_TITLE_RE = r".*"  # explorer.exe titles equal the folder name — match any

    def __init__(self, desktop: "DesktopExecutor | None" = None,
                 download_dir: str | None = None,
                 extraction_pipeline=None,
                 _subprocess_run=subprocess.run) -> None:
        self.desktop = desktop
        self.download_dir = Path(download_dir or settings.download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self._run = _subprocess_run
        self._extraction = extraction_pipeline  # injectable; built lazily

    # ── ActionRouter contract ───────────────────────────────────────────────
    def execute(self, plan: ActionPlan) -> ActionResult:
        start = time.monotonic()
        try:
            if plan.action_type == "file_navigate":
                self.open_in_explorer(plan.target)
            elif plan.action_type == "file_open":
                self.open_file(plan.target)
            elif plan.action_type == "extract_pdf":
                # plan.target = path, plan.value = comma-separated field names
                fields = [f.strip() for f in (plan.value or "").split(",") if f.strip()]
                summary = self.extract_pdf(plan.target, fields)
                return ActionResult(
                    status="ok", extracted_value=summary,
                    duration_ms=int((time.monotonic() - start) * 1000),
                )
            elif plan.action_type == "read_excel":
                rows = self.read_excel(plan.target)
                import json
                return ActionResult(
                    status="ok", extracted_value=json.dumps(rows[:10]),
                    duration_ms=int((time.monotonic() - start) * 1000),
                )
            elif plan.action_type in ("read", "extract"):
                value = self.read_text_file(plan.target)
                return ActionResult(
                    status="ok", extracted_value=value,
                    duration_ms=int((time.monotonic() - start) * 1000),
                )
            elif plan.action_type in ("flag_human", "noop"):
                pass
            else:
                return ActionResult(
                    status="failed",
                    error_msg=f"unsupported file action_type={plan.action_type!r}",
                )
            return ActionResult(
                status="ok",
                duration_ms=int((time.monotonic() - start) * 1000),
            )
        except FileError as e:
            log.warning("file_action_failed", action=plan.action_type,
                        target=plan.target, error=str(e))
            return ActionResult(status="failed", error_msg=f"file_error: {e}",
                                duration_ms=int((time.monotonic() - start) * 1000))
        except Exception as e:  # noqa: BLE001
            log.exception("file_action_crashed", action=plan.action_type)
            return ActionResult(status="failed", error_msg=f"{type(e).__name__}: {e}",
                                duration_ms=int((time.monotonic() - start) * 1000))

    # ── programmatic FS primitives ──────────────────────────────────────────
    @retry(stop=stop_after_attempt(3), wait=wait_fixed(1.0),
           retry=retry_if_exception_type(OSError), reraise=True)
    def copy_to_local(self, network_path: str) -> Path:
        """Copy a network/share file to download_dir; retries on transient SMB errors."""
        src = Path(network_path)
        if not src.exists():
            raise FileError(f"source does not exist: {network_path}")
        dst = self.download_dir / src.name
        shutil.copy2(src, dst)
        log.info("file_copy", src=str(src), dst=str(dst), size=dst.stat().st_size)
        return dst

    def find_latest_file(self, directory: str, pattern: str = "*") -> Path | None:
        dir_path = Path(directory)
        if not dir_path.exists():
            return None
        matches = sorted(dir_path.glob(pattern), key=lambda p: p.stat().st_mtime,
                         reverse=True)
        return matches[0] if matches else None

    def read_text_file(self, path: str) -> str:
        p = Path(path)
        if not p.exists():
            raise FileError(f"file does not exist: {path}")
        return p.read_text(encoding="utf-8", errors="replace")

    def acquire_lock(self, path: str) -> Path:
        """Co-operative `.lock` sentinel — caller must release_lock() after."""
        lock = Path(f"{path}.lock")
        if lock.exists():
            raise FileError(f"already locked: {lock}")
        lock.write_text(f"{os.getpid()}\n{time.time()}\n", encoding="utf-8")
        log.info("file_lock_acquired", path=str(lock))
        return lock

    def release_lock(self, path: str) -> None:
        lock = Path(f"{path}.lock")
        try:
            lock.unlink()
        except FileNotFoundError:
            return
        log.info("file_lock_released", path=str(lock))

    # ── File Explorer driving (Windows) ─────────────────────────────────────
    def open_in_explorer(self, folder_path: str) -> None:
        """Open `folder_path` in an Explorer window."""
        if sys.platform == "win32":
            self._run(["explorer.exe", folder_path], check=False)
        elif sys.platform == "darwin":
            self._run(["open", folder_path], check=False)
        else:
            self._run(["xdg-open", folder_path], check=False)
        log.info("file_navigate", folder=folder_path)
        # Verify a window exists if a desktop executor was provided.
        if self.desktop and sys.platform == "win32":
            try:
                self.desktop.attach(title_re=Path(folder_path).name)
            except Exception as e:  # noqa: BLE001
                raise FileError(f"explorer did not present the folder window: {e}")

    def open_file(self, file_path: str) -> None:
        """Open a file with the default OS handler."""
        if not Path(file_path).exists():
            raise FileError(f"file does not exist: {file_path}")
        if sys.platform == "win32":
            os.startfile(file_path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            self._run(["open", file_path], check=False)
        else:
            self._run(["xdg-open", file_path], check=False)
        log.info("file_open", path=file_path)

    # ── Excel ──────────────────────────────────────────────────────────────
    @retry(stop=stop_after_attempt(3), wait=wait_fixed(0.5),
           retry=retry_if_exception_type(OSError), reraise=True)
    def read_excel(self, path: str | Path, sheet: str | None = None,
                   header_row: int = 1) -> list[dict]:
        """Read an .xlsx into a list-of-dicts. ``header_row`` is 1-based."""
        try:
            import openpyxl
        except ImportError as e:
            raise FileError("openpyxl not installed (poetry install)") from e
        path = Path(path)
        if not path.exists():
            raise FileError(f"file does not exist: {path}")

        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        try:
            ws = wb[sheet] if sheet else wb.worksheets[0]
            rows_iter = ws.iter_rows(values_only=True)
            header = None
            out: list[dict] = []
            for idx, row in enumerate(rows_iter, start=1):
                if idx < header_row:
                    continue
                if idx == header_row:
                    header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(row)]
                    continue
                if header is None:
                    continue
                if all(c is None for c in row):
                    continue
                out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
            log.info("excel_read", path=str(path), sheet=ws.title, rows=len(out))
            return out
        finally:
            wb.close()

    def write_excel(self, path: str | Path, rows: list[dict],
                    sheet: str = "Sheet1", overwrite: bool = True) -> Path:
        """Write a list-of-dicts to an .xlsx. Atomically replaces destination."""
        try:
            import openpyxl
        except ImportError as e:
            raise FileError("openpyxl not installed (poetry install)") from e
        path = Path(path)
        if path.exists() and not overwrite:
            raise FileError(f"refusing to overwrite existing file: {path}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        if rows:
            header = list(rows[0].keys())
            ws.append(header)
            for r in rows:
                ws.append([r.get(h) for h in header])

        tmp = path.with_suffix(path.suffix + ".tmp")
        wb.save(str(tmp))
        tmp.replace(path)
        log.info("excel_write", path=str(path), rows=len(rows))
        return path

    def update_excel_cell(self, path: str | Path, sheet: str,
                          row: int, column: int | str, value) -> None:
        """1-based row / column (or column letter). Loads, edits, saves in place."""
        try:
            import openpyxl
            from openpyxl.utils import column_index_from_string
        except ImportError as e:
            raise FileError("openpyxl not installed") from e
        col_idx = column_index_from_string(column) if isinstance(column, str) else column
        wb = openpyxl.load_workbook(str(path))
        try:
            ws = wb[sheet]
            ws.cell(row=row, column=col_idx, value=value)
            wb.save(str(path))
            log.info("excel_update", path=str(path), sheet=sheet, row=row, column=column)
        finally:
            wb.close()

    # ── PDF extraction (Phase 4 pipeline) ──────────────────────────────────
    def extract_pdf(self, path: str | Path, fields: list[str] | list[dict]) -> str:
        """Run the three-tier extraction pipeline and return a JSON summary."""
        import json
        from executors.extraction import ExtractionPipeline
        if self._extraction is None:
            self._extraction = ExtractionPipeline()
        result = self._extraction.extract(path, fields)
        return json.dumps({
            "document": result.document,
            "tiers_used": result.tiers_used,
            "pages": result.pages,
            "fields": {k: v.model_dump() for k, v in result.fields.items()},
            "duration_ms": result.duration_ms,
        })

    # ── helpers used by extraction pipeline (Phase 4) ──────────────────────
    @staticmethod
    def with_temp_copy(source: str | Path):
        """Context manager: copy a (potentially shared) file to a temp local
        path, yield the path, delete on exit. Used to avoid holding network
        file handles longer than needed.
        """
        from contextlib import contextmanager

        @contextmanager
        def _cm():
            src = Path(source)
            if not src.exists():
                raise FileError(f"source does not exist: {source}")
            tmp_dir = Path(tempfile.mkdtemp(prefix="vra_"))
            local = tmp_dir / src.name
            try:
                shutil.copy2(src, local)
                yield local
            finally:
                shutil.rmtree(tmp_dir, ignore_errors=True)

        return _cm()
